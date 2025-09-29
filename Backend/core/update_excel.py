import boto3
import pandas as pd
from datetime import datetime
import os

AWS_REGION = os.getenv("AWS_REGION", "ap-south-1")
AWS_ACCESS_KEY = os.getenv("AWS_ACCESS_KEY")
AWS_SECRET_KEY = os.getenv("AWS_SECRET_KEY")
BUCKET_NAME = "ict-attendance"
EXCEL_FILE = 'students.xlsx'

s3_client = boto3.client(
    "s3",
    region_name=AWS_REGION,
    aws_access_key_id=AWS_ACCESS_KEY,
    aws_secret_access_key=AWS_SECRET_KEY,
)

from openpyxl import Workbook

def sync_students_to_excel():
    # 🔹 Get student objects from S3
    response = s3_client.list_objects_v2(Bucket=BUCKET_NAME)

    if "Contents" not in response:
        print("⚠️ No students found in S3.")
        return

    data = []
    for obj in response["Contents"]:
        key = obj["Key"]

        # Ignore the Excel file itself
        if key.endswith(".xlsx"):
            continue

        try:
            filename = os.path.basename(key) 
            batch_name = os.path.dirname(key)
            
            er_number, student_name = filename.split("_", 1)
            student_name = os.path.splitext(student_name)[0]

            last_modified = obj["LastModified"]
            upload_datetime = last_modified.strftime("%Y-%m-%d %H:%M:%S")

            data.append({
                "Batch Name": batch_name,
                "ER Number": er_number,
                "Student Name": student_name,
                "Upload Date & Time": upload_datetime,
            })

        except Exception as e:
            print(f"❌ Error parsing key {key}: {e}")

    # Create a new Workbook (fresh file every time)
    wb = Workbook()

    # Create "All Students" sheet with headers
    all_students_sheet = wb.active
    all_students_sheet.title = "All Students"
    all_students_sheet.append(["Batch Name", "ER Number", "Student Name", "Upload Date & Time"])

    # Write data to batch-specific sheets and summary
    for entry in data:
        batch_sheet_name = entry['Batch Name']

        # Create batch sheet if not exists
        if batch_sheet_name not in wb.sheetnames:
            batch_sheet = wb.create_sheet(batch_sheet_name)
            batch_sheet.append(["ER Number", "Student Name", "Upload Date & Time"])
        else:
            batch_sheet = wb[batch_sheet_name]

        # Append to batch sheet
        batch_sheet.append([
            entry['ER Number'],
            entry['Student Name'],
            entry['Upload Date & Time']
        ])

        # Append to "All Students" summary
        all_students_sheet.append([
            entry['Batch Name'],
            entry['ER Number'],
            entry['Student Name'],
            entry['Upload Date & Time']
        ])

    # Save the workbook
    wb.save(EXCEL_FILE)

    # Upload back to S3
    s3_client.upload_file(EXCEL_FILE, BUCKET_NAME, EXCEL_FILE)
    print(f"✅ Excel synced successfully with {len(data)} students.")
