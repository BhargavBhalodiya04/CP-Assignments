import os
import boto3
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from datetime import datetime
import re
import sys
from aws_config import AWS_ACCESS_KEY, AWS_SECRET_KEY, AWS_REGION
from openpyxl.utils import get_column_letter

# Constants
ALLOWED_EXTENSIONS = {'.jpg', '.jpeg', '.png'}
MAX_FILE_SIZE_MB = 5
EXCEL_FILE = 'students.xlsx'
BUCKET_NAME = 'ict-attendance'

# Initialize S3 client
s3 = boto3.client(
    's3',
    region_name=AWS_REGION,
    aws_access_key_id=AWS_ACCESS_KEY,
    aws_secret_access_key=AWS_SECRET_KEY
)


def allowed_file(filename):
    """Check allowed file extension."""
    _, ext = os.path.splitext(filename)
    return ext.lower() in ALLOWED_EXTENSIONS


def file_size_okay(file_obj):
    """Check if file is within max size."""
    file_obj.seek(0, os.SEEK_END)
    size_mb = file_obj.tell() / (1024 * 1024)
    file_obj.seek(0)
    return size_mb <= MAX_FILE_SIZE_MB


def upload_file_to_s3(bucket_name, file_path, s3_key):
    """Upload a local file to S3 with the object key."""
    try:
        print(f"Uploading to S3 → Bucket: {bucket_name}, Key: {s3_key}")
        s3.upload_file(file_path, bucket_name, s3_key)
    except Exception as e:
        raise Exception(f"Upload failed: {e}")


def sanitize_for_s3_key(text: str) -> str:
    """Remove unsafe characters and replace spaces with underscores."""
    text = text.strip().replace(" ", "_")
    text = re.sub(r'[^a-zA-Z0-9_\-]', '', text)
    return text




from openpyxl.utils import get_column_letter

def update_student_excel(batch_name, er_number, name):
    """Properly update Excel: create batch sheets and central Batch Info sheet."""
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()

    # Remove default sheet if it exists and is empty or malformed
    if 'Sheet1' in wb.sheetnames:
        sheet1 = wb['Sheet1']
        if sheet1.max_row <= 1 and sheet1.max_column <= 1:
            wb.remove(sheet1)
        elif sheet1.max_row >= 1 and sheet1.max_column >= 1 and sheet1['A1'].value != "ER Number":
            wb.remove(sheet1)

    # Ensure batch-specific sheet exists
    if batch_name in wb.sheetnames:
        batch_sheet = wb[batch_name]
    else:
        batch_sheet = wb.create_sheet(batch_name)
        batch_sheet.append(["ER Number", "Student Name", "Batch Name", "Upload Date & Time"])

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    batch_sheet.append([er_number, name, now])

    # Ensure central summary sheet exists
    summary_sheet_name = "Batch Info"
    if summary_sheet_name not in wb.sheetnames:
        summary_sheet = wb.create_sheet(summary_sheet_name)
        summary_sheet.append(["Batch Name", "ER Number", "Student Name", "Last Updated"])
    else:
        summary_sheet = wb[summary_sheet_name]

    # Avoid duplicate entries in summary
    existing_entries = [
        (row[0].value, row[1].value)
        for row in summary_sheet.iter_rows(min_row=2)
    ]
    if (batch_name, er_number) not in existing_entries:
        summary_sheet.append([batch_name, er_number, name, now])

    # Save workbook
    wb.save(EXCEL_FILE)


def upload_multiple_images(batch_name, er_number, name, image_files):
    """Upload multiple images under batch folder prefix in S3 bucket and auto-index in Rekognition."""
    er_number = er_number.strip()
    sanitized_batch_name = sanitize_for_s3_key(batch_name)
    sanitized_name = sanitize_for_s3_key(name)

    os.makedirs("uploads", exist_ok=True)
    upload_results = []

    for i, image_file in enumerate(image_files):
        filename = secure_filename(image_file.filename)
        extension = os.path.splitext(filename)[1].lower()

        if not allowed_file(filename):
            upload_results.append(f"Rejected {filename}: Invalid file type")
            continue

        if not file_size_okay(image_file):
            upload_results.append(f"Rejected {filename}: File too large (> {MAX_FILE_SIZE_MB} MB)")
            continue

        # Compose S3 key: <batch>/<er_number>_<name>_<index>.<ext>
        new_filename = f"{er_number}_{sanitized_name}_{i + 1}{extension}"
        s3_key = f"{sanitized_batch_name}/{new_filename}"
        local_path = os.path.join("uploads", new_filename)

        try:
            image_file.save(local_path)
            s3.upload_file(Filename=local_path, Bucket=BUCKET_NAME, Key=s3_key)
            upload_results.append(f"✅ Uploaded: {s3_key}")

            # 👇 Trigger Rekognition auto-index here
            index_face_to_rekognition(er_number, sanitized_name, s3_key)

        except Exception as e:
            upload_results.append(f"❌ Failed: {s3_key} -> {str(e)}")
        finally:
            if os.path.exists(local_path):
                os.remove(local_path)

    sys.dont_write_bytecode = True

    try:
        # Update Excel and upload to S3
        update_student_excel(batch_name, er_number, name)
        upload_file_to_s3(BUCKET_NAME, EXCEL_FILE, EXCEL_FILE)
        upload_results.append("✅ Excel updated and uploaded to S3 root.")
    except Exception as e:
        upload_results.append(f"❌ Excel update failed: {e}")

    return upload_results


def mark_attendance_s3():
    """
    Dummy placeholder for attendance logic using images stored in S3.
    Replace with your real face-recognition logic as needed.
    """
    results = [
        "✅ S3: Student1.jpg matched and marked present",
        "❌ S3: Student2.jpg face not found"
    ]
    return results

def index_face_to_rekognition(er_number, student_name, s3_key, collection_id="students", region="ap-south-1"):
    rekognition = boto3.client('rekognition', region_name=region)
    external_id = f"{er_number}_{student_name.replace(' ', '_')}"
    try:
        
        response = rekognition.index_faces(
            CollectionId=collection_id,
            Image={"S3Object": {"Bucket": BUCKET_NAME, "Name": s3_key}},
            ExternalImageId=external_id,
            DetectionAttributes=["DEFAULT"]
        )
        if response["FaceRecords"]:
            print(f"✅ Rekognition Indexed: {external_id}")
        else:
            print(f"⚠️ No face detected in {s3_key}")
    except rekognition.exceptions.ResourceNotFoundException:
        rekognition.create_collection(CollectionId=collection_id)
        print(f"✅ Rekognition Collection '{collection_id}' created")
        index_face_to_rekognition(er_number, student_name, s3_key, collection_id, region)


if __name__ == '__main__':
    # CLI test stub (note: no file uploads possible here)
    batch_name = input("Enter Batch Name: ").strip()
    er_number = input("Enter Student ER Number: ").strip()
    student_name = input("Enter Student Name: ").strip()
    image_files = []  # For CLI testing, must be FileStorage objects for uploads

    results = upload_multiple_images(batch_name, er_number, student_name, image_files)
    for res in results:
        print(res)
